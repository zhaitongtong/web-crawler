import PyPDF2
from openpyxl import Workbook
import openpyxl
import re
from pdfminer.pdfinterp import PDFResourceManager, PDFPageInterpreter
from pdfminer.converter import TextConverter
from pdfminer.layout import LAParams
from pdfminer.pdfpage import PDFPage
from io import StringIO


def get_text_from_pdf_2019():
    print('Getting UBC 2019 schedule of remuneration in text from PDF file ...')
    filename = 'FY19_Financial_Information_Act_Report.pdf'
    pdf_file_obj_19 = open(filename, 'rb')
    pdf_reader_19 = PyPDF2.PdfFileReader(pdf_file_obj_19)

    start_page = 40  # schedule of remuneration starts from Page 39 in the file
    end_page = 103
    count = start_page
    text = ""

    while count < end_page:
        page_obj = pdf_reader_19.getPage(count)
        count += 1
        text += page_obj.extractText()

    print('Finish getting schedule of remuneration.')
    return text


def get_text_from_pdf_2018():
    print('Getting UBC 2018 schedule of remuneration ...')
    filename = '2018-Financial-Information-Act-Report.pdf'

    rsrcmgr = PDFResourceManager()
    retstr = StringIO()
    codec = 'utf-8'
    laparams = LAParams()
    device = TextConverter(rsrcmgr, retstr, codec=codec, laparams=laparams)
    fp = open(filename, 'rb')
    interpreter = PDFPageInterpreter(rsrcmgr, device)
    password = ""
    maxpages = 0
    caching = True
    pagenos = set()

    for page in PDFPage.get_pages(fp, pagenos, maxpages=maxpages, password=password, caching=caching,
                                  check_extractable=True):
        interpreter.process_page(page)

    text = retstr.getvalue()

    fp.close()
    device.close()
    retstr.close()
    return text


def get_text_from_pdf_2017():
    print('Getting UBC 2017 schedule of remuneration ...')
    filename = 'UBC-2017_Financial-Information-Act_Report.pdf'
    pdf_file_obj_17 = open(filename, 'rb')
    pdf_reader_17 = PyPDF2.PdfFileReader(pdf_file_obj_17)

    start_page = 40
    end_page = 145
    count = start_page
    text = ""

    while count < end_page:
        page_obj = pdf_reader_17.getPage(count)
        count += 1
        text += page_obj.extractText()

    print('Finished.')
    return text


def get_salary_2019(ws):
    text_19 = get_text_from_pdf_2019()

    print('Getting salary data for year ended March 31, 2019 ...')

    curr_row = 1
    for col in ws.iter_cols(min_row=1, max_col=1, max_row=ws.max_row, values_only=True):
        for cell in col:
            full_name = ws.cell(curr_row, 3).value + ', ' + ws.cell(curr_row, 2).value
            name_index = text_19.find(full_name)
            number_index = re.search(r"\d", text_19[name_index:])
            if number_index is not None:
                space_index = text_19[name_index + number_index.start():].find(' ')
                salary = text_19[name_index + number_index.start():name_index + number_index.start() + space_index]
                prof_salary = ws.cell(row=curr_row, column=4, value=salary)
                for_year_ended = ws.cell(row=curr_row, column=5, value="March 31, 2019")
                prof_salary = ws.cell(row=curr_row, column=8, value=salary)
                for_year_ended = ws.cell(row=curr_row, column=9, value="2019")

            curr_row = curr_row + 1

    print('Finished')


def get_salary_2018(ws):
    text_18 = get_text_from_pdf_2018()
    print('Getting salary data for year ended March 31, 2018 ...')

    curr_row = 1
    for col in ws.iter_cols(min_row=1, max_col=1, max_row=ws.max_row, values_only=True):
        for cell in col:
            full_name = ws.cell(curr_row, 2).value + ' ' + ws.cell(curr_row, 3).value
            name_index = text_18.find(full_name.upper())
            if name_index < 0:
                full_name = ws.cell(curr_row, 3).value + ' ' + ws.cell(curr_row, 2).value
                name_index = text_18.find(full_name.upper())

            number_index = re.search(r"\d", text_18[name_index:])
            if number_index is not None:
                space_index = text_18[name_index + number_index.start():].find(' ')
                str = text_18[name_index + number_index.start():name_index + number_index.start() + space_index]
                comma_index = str.find(',')
                salary = str[:comma_index + 4]
                if ws.cell(curr_row, 4).value is None:
                    prof_salary = ws.cell(row=curr_row, column=4, value=salary)
                    for_year_ended = ws.cell(row=curr_row, column=5, value="March 31, 2018")
                # salary = text_18[name_index + number_index.start():name_index + number_index.start() + space_index]
                prof_salary = ws.cell(row=curr_row, column=10, value=salary)
                for_year_ended = ws.cell(row=curr_row, column=11, value="2018")

            curr_row = curr_row + 1

    print('Finished.')


def get_salary_2017(ws):
    text_17 = get_text_from_pdf_2017()
    print('Getting salary data for year ended March 31, 2017 ...')

    curr_row = 1
    for col in ws.iter_cols(min_row=1, max_col=1, max_row=ws.max_row, values_only=True):
        for cell in col:
            full_name = ws.cell(curr_row, 2).value + ' ' + ws.cell(curr_row, 3).value
            name_index = text_17.find(full_name.upper())
            if name_index < 0:
                full_name = ws.cell(curr_row, 3).value + ' ' + ws.cell(curr_row, 2).value
                name_index = text_17.find(full_name.upper())

            number_index = re.search(r"\d", text_17[name_index:])
            if number_index is not None:
                space_index = text_17[name_index + number_index.start():].find(' ')
                str = text_17[name_index + number_index.start():name_index + number_index.start() + space_index]
                comma_index = str.find(',')
                salary = str[:comma_index + 4]
                if ws.cell(curr_row, 4).value is None:
                    prof_salary = ws.cell(row=curr_row, column=4, value=salary)
                    for_year_ended = ws.cell(row=curr_row, column=5, value="March 31, 2017")

                prof_salary = ws.cell(row=curr_row, column=12, value=salary)
                for_year_ended = ws.cell(row=curr_row, column=13, value="2017")

            curr_row = curr_row + 1

    print('Finished')


def get_salary():
    wb = Workbook()
    ws = wb.active  # worksheet

    wb_prof_names = openpyxl.load_workbook('citation_University of British Columbia.xlsx')
    ws_prof_names = wb_prof_names.active

    row_count = ws_prof_names.max_row
    curr_row = 1
    for col in ws_prof_names.iter_cols(min_row=1, max_col=1, max_row=row_count, values_only=True):
        for cell in col:
            c = ws.cell(row=curr_row, column=1, value=cell)
            splitted_name = cell.split(' ')
            first_name = ws.cell(row=curr_row, column=2, value=splitted_name[0])
            last_name = ws.cell(row=curr_row, column=3, value=splitted_name[-1])
            curr_row += 1

    get_salary_2019(ws)
    get_salary_2018(ws)
    get_salary_2017(ws)
    wb.save("ProfSalaryUBC.xlsx")
    print('Finished, salary data saved to worksheet ProfSalaryUBC.xlsx')


get_salary()